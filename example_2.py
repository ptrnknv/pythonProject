import requests
import os
import zipfile
from openpyxl import load_workbook


def download_url(url, save_path, chunk_size=128):
    r = requests.get(url, stream=True)
    with open(save_path, 'wb') as fd:
        for chunk in r.iter_content(chunk_size=chunk_size):
            fd.write(chunk)


path_zip = '/home/petr/Desktop/PycharmProjects/pythonProject/task.zip'
path_unzip = '/home/petr/Desktop/PycharmProjects/pythonProject/task/'
if not os.path.isfile(path_zip):
    url = 'https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip'
    download_url(url, path_zip)

if not os.path.isfile(path_unzip):
    with zipfile.ZipFile(path_zip, "r") as zip_ref:
        zip_ref.extractall(path_unzip)

directory = os.fsencode(path_unzip)
workers = []
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    workbook = load_workbook(filename=path_unzip + filename)
    sheet = workbook.active
    for value in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
        if value[0] == 'ФИО':
            workers.append(value)

with open('result.txt', 'w', encoding='utf8') as file:
    for worker in sorted(workers, key=lambda x: x[1]):
        file.write(f'{worker[1]} {worker[3]}\n')
